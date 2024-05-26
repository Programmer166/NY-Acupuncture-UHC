import EOB
import os
import FromCsvToIncome
# import pandas as pd
import xlsxwriter
import csv
import openpyxl

# EOB_FILE_PATH=r"D:\EOB\eob_path"
EOB_FILE_PATH = os.getcwd() + r"\EOB"


def main():
    EOB.running(EOB_FILE_PATH)


# print(os.path.exists(EOB_FILE_PATH))
main()

if os.path.exists('EOB.xlsx'):
    os.remove('EOB.xlsx')

if os.path.exists('faillist.xlsx'):
    os.remove('faillist.xlsx')

if os.path.exists('faillist_new.xlsx'):
    os.remove('faillist_new.xlsx')

# read_file = pd.read_csv (r'EOB.csv', header=None)
# read_file.to_excel (r'EOB.xlsx', index = None, header=None)
wb_test = openpyxl.Workbook()
ws_test = wb_test.active
if os.path.exists('EOB.csv'):
    with open('EOB.csv') as f_test:
        reader_test = csv.reader(f_test)  # comma default for me
        for row in reader_test:
            ws_test.append(row)
    wb_test.save('EOB.xlsx')

FromCsvToIncome.receive_fees()
print("..................")
print("This is NewIncome......")
print("..................")
FromCsvToIncome.receive_fees_new()

fail_fp = xlsxwriter.Workbook('faillist.xlsx')
fail_fp_sheet = fail_fp.add_worksheet()

row = 0
column = 0

print("This is the failed in Income......")
print("The total number of the failed in Income is ", len(FromCsvToIncome.fail_list))

# iterating through content list
for fail_item in FromCsvToIncome.fail_list:
    # write operation perform
    print(fail_item)
    fail_item = str(fail_item).split("*")
    fail_fp_sheet.write(row, 1, fail_item[0])
    fail_fp_sheet.write(row, 2, fail_item[1])
    fail_fp_sheet.write(row, 0, fail_item[2])
    fail_fp_sheet.write(row, 8, fail_item[3])
    fail_fp_sheet.write(row, 4, fail_item[4])
    fail_fp_sheet.write(row, 20, fail_item[5])
    fail_fp_sheet.write(row, 6, fail_item[6])
    fail_fp_sheet.write(row, 7, fail_item[7])
    # print(fail_item)

    # incrementing the value of row by one
    # with each iterations.
    row += 1

fail_fp.close()

fail_fp_new = xlsxwriter.Workbook('faillist_new.xlsx')
fail_fp_sheet_new = fail_fp_new.add_worksheet()

row_new = 0
column_new = 0

print("This is the failed in NewIncome......")
print("The total number of the failed in NewIncome is ", len(FromCsvToIncome.fail_list_new))
# iterating through content list
for fail_item_new in FromCsvToIncome.fail_list_new:
    # write operation perform
    # fail_fp_sheet_new.write(row_new, column_new, fail_item_new)
    print(fail_item_new)
    fail_item_new = str(fail_item_new).split("*")
    fail_fp_sheet_new.write(row_new, 1, fail_item_new[0])
    fail_fp_sheet_new.write(row_new, 2, fail_item_new[1])
    fail_fp_sheet_new.write(row_new, 0, fail_item_new[2])
    fail_fp_sheet_new.write(row_new, 8, fail_item_new[3])
    fail_fp_sheet_new.write(row_new, 4, fail_item_new[4])
    fail_fp_sheet_new.write(row_new, 20, fail_item_new[5])
    fail_fp_sheet_new.write(row_new, 6, fail_item_new[6])
    fail_fp_sheet_new.write(row_new, 7, fail_item_new[7])

    # incrementing the value of row by one
    # with each iterations.
    row_new += 1

fail_fp_new.close()

print("Successful......")
while 1:
    pass

# E:\Code\Jobs\NewYorkAcupunctureChiropractic\EOBReceiveFees\EOB\eob_path


