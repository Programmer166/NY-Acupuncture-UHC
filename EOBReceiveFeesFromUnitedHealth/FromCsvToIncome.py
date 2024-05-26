# importing pandas module

import os
import openpyxl
from datetime import date
import re




today = date.today()
#ACCOUNTING_PATH = r"C:\Users\Evelyn\Google Drive\income"
#ACCOUNTING_PATH_NEW = r"C:\Users\Evelyn\Google Drive\income\NewIncome"
ACCOUNTING_PATH = r"C:\Users\demki\Desktop\AcupuntureMaterials\income"
ACCOUNTING_PATH_NEW = r"C:\Users\demki\Desktop\AcupuntureMaterials\income\NewIncome"
fail_list = []
fail_list_new = []



def search_in_patient(filename, last_name, first_name):
    if bool(re.search(str(last_name), filename, re.IGNORECASE) and re.search(str(first_name), filename,
                                                                        re.IGNORECASE) and re.search(".xlsx", filename,
                                                                                                     re.IGNORECASE) and re.search(
            "income", filename, re.IGNORECASE)):


        return 1
    else:
        return 0





def receive_fees():
    current_directory = os.getcwd()
    #print(os.getcwd())
    #print(current_directory)
    #print(type(current_directory))

    #fp = os.path.join(os.getcwd(), r"\EOB.xlsx")
    fp = os.path.join(current_directory, r"EOB.xlsx")
    insurance_list = openpyxl.load_workbook(fp)
    insurance_table = insurance_list.active
    temp_num = 0

    for i in range(1, 1 + insurance_table.max_row):

        temp_num = temp_num + 1

        tempRow = []
        for tempi in range(1, 18):
            tempCellValue = insurance_table.cell(i, tempi).value
            tempRow.append(tempCellValue)

        last_name=tempRow[3]
        first_name=tempRow[4]
        paid_amount=tempRow[7].strip(" ")
        trace_amount_num = tempRow[13]
        #print(paid_amount)
        #print(type(paid_amount))
        charge_amount=str(tempRow[6]).replace("$", "")
        today_date=str(today.strftime("%m/%d/%Y"))
        #print("new claim")
        #today_date_test = "test" + today_date
        #print(today_date_test)
        insurance_id=tempRow[0]
        service_date=tempRow[5]
        #print(service_date)
        temp_service_date = str(service_date).split("/", str(service_date).count("/"))
        if len(temp_service_date) != 3:
            fail_info = str(last_name) + "*" + str(first_name) + "*" + str(service_date) + "*" + str(paid_amount) + "*" + str(charge_amount) + "*" + str(trace_amount_num) + "*" + "UNITED HEALTHCARE" + "*" + today_date
            fail_list.append(fail_info)
            continue
        temp_service_yearfour = temp_service_date[2].lstrip('0')
        temp_length = len(temp_service_yearfour)
        temp_first_position = temp_length - 2
        temp_service_yeartwo = temp_service_yearfour[temp_first_position:]
        #new_service_date = temp_service_date[0].lstrip('0') + temp_service_date[1].lstrip('0') + temp_service_date[2].lstrip('0')
        new_service_date = temp_service_date[0].lstrip('0') + temp_service_date[1].lstrip('0') + temp_service_yeartwo
        temp_charge_amount = charge_amount.split(".", charge_amount.count("."))
        new_charge_amount = temp_charge_amount[0]
        #print(new_charge_amount)
        #print(new_service_date)

        temp_filesearch_result = 0
        for filename in os.listdir(ACCOUNTING_PATH):
            #temp_filesearch_result=0

            patient_fp = os.path.join(ACCOUNTING_PATH, filename)
            if os.path.isfile(patient_fp):
                search_result=search_in_patient(filename, last_name, first_name)
                if search_result == 1:
                    #print(filename)
                    #print(str(patient_fp))
                    patientWb = openpyxl.load_workbook(patient_fp)
                    name_list = patientWb.get_sheet_names()
                    my_sheet = patientWb.get_sheet_by_name(name_list[0])
                    for templ in range(1, 1 + my_sheet.max_row):
                        patient_service_date = str(my_sheet.cell(templ, 1).value)
                        #print(patient_service_date)
                        #print(templ,"-----------------------------------------")
                        if patient_service_date.count("/") == 2:
                            #print(patient_service_date)
                            temp_patient_date = patient_service_date.split("/", patient_service_date.count("/"))
                            year_four = temp_patient_date[2].lstrip('0')
                            length = len(year_four)
                            first_position = length - 2
                            year_two = year_four[first_position:]
                            #print(year_two)

                            new_patient_date = temp_patient_date[0].lstrip('0') + temp_patient_date[1].lstrip('0') + year_two
                            patient_charge_amount = str(my_sheet.cell(templ, 5).value)
                            temp1_patient_amount = patient_charge_amount.lstrip('$')
                            temp2_patient_amount = temp1_patient_amount.split(".", temp1_patient_amount.count("."))
                            new_patient_amount = temp2_patient_amount[0]
                            #print(new_patient_amount)
                            #print(new_patient_date)

                            if new_service_date == new_patient_date and new_charge_amount == new_patient_amount:
                                if paid_amount != "$0.00":
                                    my_sheet.cell(templ, 8).value = today_date
                                    my_sheet.cell(templ, 9).value = paid_amount
                                    my_sheet.cell(templ, 21).value = trace_amount_num

                                #print("time is matched and the charge amount is also matched")

                                #my_sheet.cell(templ, 9).value = paid_amount
                                temp_filesearch_result = 1
                                break

                        elif patient_service_date.count("-") == 2:
                            #print(patient_service_date)
                            temp_patient_date = patient_service_date.split("-", patient_service_date.count("-"))
                            year_four = temp_patient_date[0].lstrip('0')
                            length = len(year_four)
                            first_position = length - 2
                            year_two = year_four[first_position:]
                            #print(year_two)

                            temp_service_day = temp_patient_date[2].split(" ", patient_service_date.count(" "))
                            service_day = temp_service_day[0].lstrip('0')

                            new_patient_date = temp_patient_date[1].lstrip('0') + service_day + year_two
                            patient_charge_amount = str(my_sheet.cell(templ, 5).value)
                            temp1_patient_amount = patient_charge_amount.lstrip('$')
                            temp2_patient_amount = temp1_patient_amount.split(".", temp1_patient_amount.count("."))
                            new_patient_amount = temp2_patient_amount[0]
                            #print(new_patient_amount)
                            #print(new_patient_date)

                            if new_service_date == new_patient_date and new_charge_amount == new_patient_amount:
                                if paid_amount != "$0.00":
                                    my_sheet.cell(templ, 8).value = today_date
                                    my_sheet.cell(templ, 9).value = paid_amount
                                    my_sheet.cell(templ, 21).value = trace_amount_num
                                #print("time is matched and the charge amount is also matched")

                                #my_sheet.cell(templ, 9).value = paid_amount
                                temp_filesearch_result = 1
                                break

                        else:
                            #print(patient_service_date)
                            print("date formate is not correct")

                    patientWb.save(patient_fp)


        if temp_filesearch_result == 1:
            # print(last_name+"/" + first_name + "/" +service_date + "/" + insurance_id)
            print(temp_num)
            #break
        else:
            fail_info = str(last_name) + "*" + str(first_name) + "*" + str(service_date) + "*" + str(paid_amount) + "*" + str(charge_amount) + "*" + str(trace_amount_num) + "*" + "UNITED HEALTHCARE" + "*" + today_date
            fail_list.append(fail_info)











def receive_fees_new():
    current_directory = os.getcwd()
    #print(os.getcwd())
    #print(current_directory)
    #print(type(current_directory))

    #fp = os.path.join(os.getcwd(), r"\EOB.xlsx")
    fp = os.path.join(current_directory, r"EOB.xlsx")
    insurance_list = openpyxl.load_workbook(fp)
    insurance_table = insurance_list.active
    temp_num = 0

    for i in range(1, 1 + insurance_table.max_row):

        temp_num = temp_num + 1

        tempRow = []
        for tempi in range(1, 18):
            tempCellValue = insurance_table.cell(i, tempi).value
            tempRow.append(tempCellValue)

        last_name=tempRow[3]
        first_name=tempRow[4]
        paid_amount=tempRow[7]
        trace_amount_num = tempRow[13]
        charge_amount=str(tempRow[6])
        today_date=str(today.strftime("%m/%d/%Y"))
        #print("new claim")
        #today_date_test = "test" + today_date
        #print(today_date_test)
        insurance_id=tempRow[0]
        service_date=tempRow[5]
        #print(service_date)
        temp_service_date = str(service_date).split("/", str(service_date).count("/"))
        if len(temp_service_date) != 3:
            fail_info = str(last_name) + "*" + str(first_name) + "*" + str(service_date) + "*" + str(paid_amount) + "*" + str(charge_amount) + "*" + str(trace_amount_num) + "*" + "UNITED HEALTHCARE" + "*" + today_date
            fail_list_new.append(fail_info)
            continue
        temp_service_yearfour = temp_service_date[2].lstrip('0')
        temp_length = len(temp_service_yearfour)
        temp_first_position = temp_length - 2
        temp_service_yeartwo = temp_service_yearfour[temp_first_position:]
        # new_service_date = temp_service_date[0].lstrip('0') + temp_service_date[1].lstrip('0') + temp_service_date[2].lstrip('0')
        new_service_date = temp_service_date[0].lstrip('0') + temp_service_date[1].lstrip('0') + temp_service_yeartwo
       # new_service_date = temp_service_date[0].lstrip('0') + temp_service_date[1].lstrip('0') + temp_service_date[2].lstrip('0')
        temp_charge_amount = charge_amount.split(".", charge_amount.count("."))
        new_charge_amount = temp_charge_amount[0]
        #print(new_charge_amount)
        #print(new_service_date)

        temp_filesearch_result = 0
        for filename in os.listdir(ACCOUNTING_PATH_NEW):
            #temp_filesearch_result=0

            patient_fp = os.path.join(ACCOUNTING_PATH_NEW, filename)
            if os.path.isfile(patient_fp):
                search_result=search_in_patient(filename, last_name, first_name)
                if search_result == 1:
                    #print(filename)
                    #print(str(patient_fp))
                    patientWb = openpyxl.load_workbook(patient_fp)
                    name_list = patientWb.get_sheet_names()
                    my_sheet = patientWb.get_sheet_by_name(name_list[0])
                    for templ in range(1, 1 + my_sheet.max_row):
                        patient_service_date = str(my_sheet.cell(templ, 1).value)
                        #print(patient_service_date)
                        if patient_service_date.count("/") == 2:
                            temp_patient_date = patient_service_date.split("/", patient_service_date.count("/"))
                            year_four = temp_patient_date[2].lstrip('0')
                            length = len(year_four)
                            first_position = length - 2
                            year_two = year_four[first_position:]
                            #print(year_two)

                            new_patient_date = temp_patient_date[0].lstrip('0') + temp_patient_date[1].lstrip('0') + year_two
                            patient_charge_amount = str(my_sheet.cell(templ, 5).value)
                            temp1_patient_amount = patient_charge_amount.lstrip('$')
                            temp2_patient_amount = temp1_patient_amount.split(".", temp1_patient_amount.count("."))
                            new_patient_amount = temp2_patient_amount[0]
                            #print(new_patient_amount)
                            #print(new_patient_date)

                            if new_service_date == new_patient_date and new_charge_amount == new_patient_amount:
                                if paid_amount != "$0.00":
                                    my_sheet.cell(templ, 8).value = today_date
                                    my_sheet.cell(templ, 9).value = paid_amount
                                    my_sheet.cell(templ, 21).value = trace_amount_num


                                #my_sheet.cell(templ, 9).value = paid_amount
                                temp_filesearch_result = 1
                                break

                        elif patient_service_date.count("-") == 2:
                            temp_patient_date = patient_service_date.split("-", patient_service_date.count("-"))
                            year_four = temp_patient_date[0].lstrip('0')
                            length = len(year_four)
                            first_position = length - 2
                            year_two = year_four[first_position:]
                            #print(year_two)

                            temp_service_day = temp_patient_date[2].split(" ", patient_service_date.count(" "))
                            service_day = temp_service_day[0].lstrip('0')

                            new_patient_date = temp_patient_date[1].lstrip('0') + service_day + year_two
                            patient_charge_amount = str(my_sheet.cell(templ, 5).value)
                            temp1_patient_amount = patient_charge_amount.lstrip('$')
                            temp2_patient_amount = temp1_patient_amount.split(".", temp1_patient_amount.count("."))
                            new_patient_amount = temp2_patient_amount[0]
                            #print(new_patient_amount)
                            #print(new_patient_date)

                            if new_service_date == new_patient_date and new_charge_amount == new_patient_amount:
                                if paid_amount != "$0.00":
                                    my_sheet.cell(templ, 8).value = today_date
                                    my_sheet.cell(templ, 9).value = paid_amount
                                    my_sheet.cell(templ, 21).value = trace_amount_num


                                #my_sheet.cell(templ, 9).value = paid_amount
                                temp_filesearch_result = 1
                                break

                        else:
                            print("date formate is not correct")

                    patientWb.save(patient_fp)


        if temp_filesearch_result == 1:
            # print(last_name+"/" + first_name + "/" +service_date + "/" + insurance_id)
            print(temp_num)
            #break
        else:
            fail_info = str(last_name) + "*" + str(first_name) + "*" + str(service_date) + "*" + str(paid_amount) + "*" + str(charge_amount) + "*" + str(trace_amount_num) + "*" + "UNITED HEALTHCARE" + "*" + today_date
            fail_list_new.append(fail_info)









'''
def receive_fees_new():
    current_directory = os.getcwd()
    #print(os.getcwd())
    #print(current_directory)
    #print(type(current_directory))

    #fp = os.path.join(os.getcwd(), r"\EOB.xlsx")
    fp = os.path.join(current_directory, r"EOB.xlsx")
    insurance_list = openpyxl.load_workbook(fp)
    insurance_table = insurance_list.active
    temp_num_new = 0

    for i in range(1, 1 + insurance_table.max_row):
        temp_num_new = temp_num_new + 1

        tempRow = []
        for tempi in range(1, 18):
            tempCellValue = insurance_table.cell(i, tempi).value
            tempRow.append(tempCellValue)

        last_name=tempRow[3]
        first_name=tempRow[4]
        paid_amount=tempRow[7]
        charge_amount=str(tempRow[6])
        today_date=str(today.strftime("%m/%d/%Y"))
        #print("new claim")
        #today_date_test = "test" + today_date
        #print(today_date_test)
        insurance_id=tempRow[0]
        service_date=tempRow[5]
        #print(service_date)
        temp_service_date = str(service_date).split("/", str(service_date).count("/"))
        if len(temp_service_date) != 3:
            continue
        new_service_date = temp_service_date[0].lstrip('0') + temp_service_date[1].lstrip('0') + temp_service_date[2].lstrip('0')
        temp_charge_amount = charge_amount.split(".", charge_amount.count("."))
        new_charge_amount = temp_charge_amount[0]
        #print(new_charge_amount)
        #print(new_service_date)


        for filename in os.listdir(ACCOUNTING_PATH_NEW):
            temp_filesearch_result=0

            patient_fp = os.path.join(ACCOUNTING_PATH_NEW, filename)
            if os.path.isfile(patient_fp):
                search_result=search_in_patient(filename, last_name, first_name)
                if search_result == 1:
                    #print(filename)
                    #print(str(patient_fp))
                    patientWb = openpyxl.load_workbook(patient_fp)
                    name_list = patientWb.get_sheet_names()
                    my_sheet = patientWb.get_sheet_by_name(name_list[0])
                    for templ in range(1, 1 + my_sheet.max_row):
                        patient_service_date = str(my_sheet.cell(templ, 1).value)
                        #print(patient_service_date)
                        if patient_service_date.count("/") == 2:
                            temp_patient_date = patient_service_date.split("/", patient_service_date.count("/"))
                            year_four = temp_patient_date[2].lstrip('0')
                            length = len(year_four)
                            first_position = length - 2
                            year_two = year_four[first_position:]
                            #print(year_two)

                            new_patient_date = temp_patient_date[0].lstrip('0') + temp_patient_date[1].lstrip('0') + year_two
                            patient_charge_amount = str(my_sheet.cell(templ, 5).value)
                            temp1_patient_amount = patient_charge_amount.lstrip('$')
                            temp2_patient_amount = temp1_patient_amount.split(".", temp1_patient_amount.count("."))
                            new_patient_amount = temp2_patient_amount[0]
                            #print(new_patient_amount)
                            #print(new_patient_date)

                            if new_service_date == new_patient_date and new_charge_amount == new_patient_amount:
                                my_sheet.cell(templ, 8).value = today_date
                                my_sheet.cell(templ, 9).value = paid_amount
                                temp_filesearch_result = 1
                                break

                        elif patient_service_date.count("-") == 2:
                            temp_patient_date = patient_service_date.split("-", patient_service_date.count("-"))
                            year_four = temp_patient_date[0].lstrip('0')
                            length = len(year_four)
                            first_position = length - 2
                            year_two = year_four[first_position:]
                            #print(year_two)

                            temp_service_day = temp_patient_date[2].split(" ", patient_service_date.count(" "))
                            service_day = temp_service_day[0].lstrip('0')

                            new_patient_date = temp_patient_date[1].lstrip('0') + service_day + year_two
                            patient_charge_amount = str(my_sheet.cell(templ, 5).value)
                            temp1_patient_amount = patient_charge_amount.lstrip('$')
                            temp2_patient_amount = temp1_patient_amount.split(".", temp1_patient_amount.count("."))
                            new_patient_amount = temp2_patient_amount[0]
                            #print(new_patient_amount)
                            #print(new_patient_date)

                            if new_service_date == new_patient_date and new_charge_amount == new_patient_amount:
                                my_sheet.cell(templ, 8).value = today_date
                                my_sheet.cell(templ, 9).value = paid_amount
                                temp_filesearch_result = 1
                                break

                        else:
                            print("date formate is not correct")








                    patientWb.save(patient_fp)
                    if temp_filesearch_result==1:
                        #print(last_name + "/" + first_name + "/" + service_date + "/" + insurance_id)
                        print(temp_num_new)
                        break
                    else:
                        fail_info = last_name+"/" + first_name + "/" +service_date + "/" + insurance_id
                        fail_list_new.append(fail_info)
'''






